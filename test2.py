import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import os
import openpyxl as xl
import openpyxl.utils.cell as xlutils
import pyexcel

def evaluate_formula(cell, sheet):
    formula = cell.value
    evaluated_value = cell.internal_value
    if cell.data_type == "f":
        book = pyexcel.get_book(file_name=sheet.parent.parent.path)
        sheet_index = sheet.parent.index(sheet)
        evaluated_value = book[sheet_index].at(cell.row - 1, cell.col - 1)

    return evaluated_value

# Inisiasi UI
windows = tk.Tk()
windows.title("Aplikasi pencari data berdasarkan NIK")
windows.geometry("480x360")
windows.resizable(False, False)

# Deklarasi global variabel
workbook = None
sheet = None
text_error = None
text2 = None
text3 = None
text4 = None
text_tgl = None
text_loc = None

# Function untuk membuka file Excel
def membuka_file():
    global workbook, sheet, sheet_selector, text2, text3, text4, text_tgl, text_loc
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        try:
            workbook = xl.load_workbook(file_path)
            sheet = workbook.active
            sheet_names = workbook.sheetnames
            sheet_selector['values'] = sheet_names  # Update the dropdown menu with sheet names
            sheet_selector.current(0)  # Select the first sheet by default
            judul1.config(text="Nama file: " + os.path.basename(file_path))
            sheet_selector.place(x=25, y=328)  # Place the selector in the bottom left
            sheet_selector.config(state="readonly")  # Enable the combobox
            clear_results()  # Clear previous search results
            return True
        except:
            return False
        
# Function to handle sheet selection from the dropdown menu
def select_sheet(event):
    global sheet
    selected_sheet = sheet_selector.get()
    sheet = workbook[selected_sheet]


# Function untuk mengecek apakah file berhasil dibuka
def tombol():
    if membuka_file():
        tulisan = tk.Label(windows, text="File terbuka.\t\t\t", font=("Arial", 10))
        tulisan.place(x=75, y=60)
        info1.config(text="Status: File terbuka")
        tombol_klik1.config(state="normal")
        nomorNik.config(state="normal")
        tombol_klik.place(x=380, y=28)  # Move the button to the right
    else:
        tulisan = tk.Label(windows, text="Gagal untuk membuka file!!", font=("Arial", 10), foreground='red')
        tulisan.place(x=75, y=60)
        info1.config(text="Status: Gagal membuka file")
        tombol_klik.place(x=208, y=28)  # Move the button to the center

# Function untuk mengecek apakah NIK yang diinputkan ada di file Excel
def pengecekan_nik():
    global sheet, text_error, text2, text3, text4, text_tgl, text_loc
    nik = nomorNik.get()

    # Clear previous search results
    clear_results()

    row_labels = [cell.value for cell in sheet[1]]  # Get the row labels from the first row

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[2]) == str(nik):
            for i, value in enumerate(row, start=0):
                cell = sheet.cell(row=row_num, column=i+1)
                evaluated_value = evaluate_formula(cell, sheet)

                label_text = f"{row_labels[i]}: {evaluated_value}"
                label = tk.Label(windows, text=label_text)
                label.place(x=25, y=150 + (20 * i))

            text_loc = tk.Label(windows, text="Lokasi data berada di baris ke " + str(row_num))
            text_loc.place(x=25, y=290)

            try:
                text_error.destroy()
            except:
                text_error = None
            break
    else:
        text_error = tk.Label(windows, text="Data dengan NIK " + str(nik) + " tidak ditemukan!!", foreground='red')
        text_error.place(x=25, y=300)

    # Reset the NIK input field
    nomorNik.delete(0, tk.END)



# Function to clear previous search results
def clear_results():
    global text2, text3, text4, text_error, text_tgl, text_loc
    if text2:
        text2.destroy()
    if text3:
        text3.destroy()
    if text4:
        text4.destroy()
    if text_error:
        text_error.destroy()
    if text_tgl:
        text_tgl.destroy()
    if text_loc:
        text_loc.destroy()

# UI elements
garis = tk.Canvas(windows, width=500)
garis.place(x=14, y=-88)
garis.create_line(0, 180, 447, 180)
garis.create_line(0, 230, 180, 230)
garis.create_line(250, 230, 447, 230)

label_pack = tk.Label(windows, text="Aplikasi by kelAI", font=("Arial", 9))
label_pack.place(x=355, y=338)

judul1 = tk.Label(windows, font=("Arial", 11))
judul1.place(x=25, y=30)

judul2 = tk.Label(windows, text="Masukan no NIK : ", font=("Arial", 11))
judul2.place(x=25, y=105)

judul3 = tk.Label(windows, text="Info Data")
judul3.place(x=205, y=130)

info1 = tk.Label(windows, font=("Arial", 10))
info1.place(x=25, y=60)

nomorNik = tk.Entry(windows, width=25, state="disabled")
nomorNik.place(x=164, y=108)

tombol_klik = tk.Button(windows, text="Buka File", command=tombol, width=12, height=2)  # Increase the button size
tombol_klik.place(x=197, y=28)  # Initially place the button in the center

tombol_klik1 = tk.Button(windows, text="Cari NIK", command=pengecekan_nik, state="disabled")
tombol_klik1.place(x=330, y=103)

sheet_selector = ttk.Combobox(windows, state="enabled")
sheet_selector.bind("<<ComboboxSelected>>", select_sheet)
windows.mainloop()
